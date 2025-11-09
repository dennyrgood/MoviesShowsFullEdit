import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import shutil # Required for file copying (backup)
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table
from datetime import datetime, date
import sys

class ExcelRowInserterApp:
    """
    A Tkinter application to insert a new row into an Excel spreadsheet.
    The required backup of the target file is now performed unconditionally 
    at the time the application loads, before the GUI is displayed.
    """
    def __init__(self, master):
        self.master = master
        master.title("Movie/Show Data Inserter (Table-Aware)")
        master.geometry("650x850") # Adjusted geometry to fit the new legend tables

        # Default File Paths
        self.DEFAULT_TARGET_FILE = "Movies and Shows.xlsx" # Single target file path

        # --- Initial Backup on Application Load (BEFORE GUI RENDERS) ---
        target_file = self.DEFAULT_TARGET_FILE
        backup_success, backup_info = self.create_backup(target_file)
        
        if not backup_success:
            if "Original file not found" in backup_info:
                # If file doesn't exist, show error and proceed (user might select it later)
                messagebox.showerror("Initial Backup Error", f"Target file not found at:\n{target_file}\n\nInitial backup skipped. Please ensure the file exists or select the correct path.")
            else:
                # Other errors (e.g., permission denied)
                messagebox.showwarning("Initial Backup Warning", f"Could not create initial backup of the file at '{target_file}':\n{backup_info}\n\nProceeding without backup. Use caution.")
            self.initial_backup_status = "FAILED"
            self.initial_backup_message = backup_info
        else:
            # REMOVED: messagebox.showinfo for successful backup. Screen status is sufficient.
            self.initial_backup_status = "SUCCESS"
            self.initial_backup_message = backup_info


        # Style Configuration
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TFrame', background='#e8f5e9')
        style.configure('TLabel', background='#e8f5e9', font=('Arial', 10))
        style.configure('TButton', font=('Arial', 10, 'bold'), background='#4caf50', foreground='white')
        style.map('TButton', background=[('active', '#66bb6a')])
        style.configure('Success.TLabel', foreground='#2e7d32', font=('Arial', 11, 'bold'))
        style.configure('Error.TLabel', foreground='#c62828', font=('Arial', 11, 'bold'))
        style.configure('TEntry', fieldbackground='white')

        # Custom style for legend table
        style.configure('Legend.TLabel', background='#f1f8e9', font=('Courier', 9), padding=2)
        style.configure('LegendHeader.TLabel', background='#dcedc8', font=('Courier', 9, 'bold'), padding=3)

        # --- Main Frame ---
        main_frame = ttk.Frame(master, padding="25")
        main_frame.pack(fill='both', expand=True)

        # Title
        title_label = ttk.Label(main_frame, text="Insert New Entry (Movies/Shows) Into Table", font=('Arial', 16, 'bold'), background='#c8e6c9', padding=10)
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 25), sticky='ew')
        
        # --- File Selection Section ---
        backup_text = "Target File Path (Initial backup completed on load)" if self.initial_backup_status == "SUCCESS" else "Target File Path (Initial backup FAILED - Check path/permissions)"
        file_frame = ttk.LabelFrame(main_frame, text=backup_text, padding="15", borderwidth=2, relief="groove")
        file_frame.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(10, 20))

        # Target File (combines old Input and Output)
        ttk.Label(file_frame, text="Target Excel File (Required):").grid(row=0, column=0, sticky='w', pady=5, padx=5)
        self.target_file_path = tk.StringVar(value=self.DEFAULT_TARGET_FILE)
        self.target_entry = ttk.Entry(file_frame, textvariable=self.target_file_path, width=40)
        self.target_entry.grid(row=0, column=1, sticky='ew', padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_target_file).grid(row=0, column=2, padx=5)

        file_frame.grid_columnconfigure(1, weight=1)

        # --- Data Entry Section ---
        data_frame = ttk.LabelFrame(main_frame, text="New Row Data Entry (A, F, G, H, I, J, K, L, M)", padding="15", borderwidth=2, relief="groove")
        data_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=10)

        # Data fields definition
        
        # Initialize StringVars
        self.col_a = tk.StringVar()
        self.col_f = tk.StringVar()
        self.col_g = tk.StringVar()
        self.col_h = tk.StringVar()
        self.col_i = tk.StringVar(value="Download") # Default value
        self.col_j = tk.StringVar()
        self.col_k = tk.StringVar()
        self.col_l = tk.StringVar()
        self.col_m = tk.StringVar()

        fields = [
            # Code field (Number/Decimal)
            ("A: Input Code (MANDATORY, Must be a Number/Decimal):", self.col_a), 
            ("F: Title (MANDATORY):", self.col_f),
            ("G: First Season/Get (Optional):", self.col_g),
            ("H: To Watch/Where (Optional):", self.col_h),
            ("I: Seasons/Made (Optional, default 'Download'):", self.col_i),
            # Date field (Column J) - Relaxed input format, saved as DD-MMM-YYYY
            ("J: Season/Data Avail (Optional, e.g., 1/1/2025 or 10-Oct-2025):", self.col_j), 
            ("K: Website/Notes (Optional):", self.col_k),
            ("L: Notes (Optional):", self.col_l),
            ("M: Synopsis (Optional):", self.col_m),
        ]

        for i, (label_text, var) in enumerate(fields):
            ttk.Label(data_frame, text=label_text).grid(row=i, column=0, sticky='w', pady=4, padx=5)
            entry = ttk.Entry(data_frame, textvariable=var, width=60)
            entry.grid(row=i, column=1, sticky='ew', padx=5)
            
        data_frame.grid_columnconfigure(1, weight=1)

        # --- Code Legend Section (New) ---
        legend_frame = ttk.LabelFrame(main_frame, text="Code (Column A) Legend (Whole Number . Decimal)", padding="15", borderwidth=2, relief="groove")
        legend_frame.grid(row=3, column=0, columnspan=2, sticky='ew', pady=10)
        
        # Whole Number Legend Table (Who is interested)
        ttk.Label(legend_frame, text="Whole Number (Who)", font=('Arial', 10, 'bold'), anchor='center').grid(row=0, column=0, columnspan=2, pady=(0, 5), sticky='ew')
        
        ttk.Label(legend_frame, text="Code Range", style='LegendHeader.TLabel').grid(row=1, column=0, sticky='ew')
        ttk.Label(legend_frame, text="Meaning", style='LegendHeader.TLabel').grid(row=1, column=1, sticky='ew')
        
        ttk.Label(legend_frame, text="0.1 - 0.9", style='Legend.TLabel').grid(row=2, column=0, sticky='ew')
        ttk.Label(legend_frame, text="Laura", style='Legend.TLabel').grid(row=2, column=1, sticky='ew')

        ttk.Label(legend_frame, text="1.1 - 1.9", style='Legend.TLabel').grid(row=3, column=0, sticky='ew')
        ttk.Label(legend_frame, text="Laura Maybe", style='Legend.TLabel').grid(row=3, column=1, sticky='ew')

        ttk.Label(legend_frame, text="2.1 - 2.9", style='Legend.TLabel').grid(row=4, column=0, sticky='ew')
        ttk.Label(legend_frame, text="Not Laura", style='Legend.TLabel').grid(row=4, column=1, sticky='ew')

        ttk.Label(legend_frame, text="1000.x", style='Legend.TLabel').grid(row=5, column=0, sticky='ew')
        ttk.Label(legend_frame, text="Laura Only", style='Legend.TLabel').grid(row=5, column=1, sticky='ew')
        
        # Spacer
        ttk.Frame(legend_frame, height=10, relief='flat').grid(row=6, column=0, columnspan=2)

        # Decimal Legend Table (Acquisition Status)
        ttk.Label(legend_frame, text="Decimal Part (Acquisition Status)", font=('Arial', 10, 'bold'), anchor='center').grid(row=7, column=0, columnspan=2, pady=(5, 5), sticky='ew')

        ttk.Label(legend_frame, text="Decimal Code", style='LegendHeader.TLabel').grid(row=8, column=0, sticky='ew')
        ttk.Label(legend_frame, text="Meaning", style='LegendHeader.TLabel').grid(row=8, column=1, sticky='ew')

        decimal_data = [
            (".1", "Avail Now"),
            (".4", "Weekly"),
            (".5", "Trying to Find"),
            (".6", "Found but $ (Paid)"),
            (".7", "Future Date Set"),
            (".8", "Announced as Coming"),
            (".9", "Unknown if Coming"),
        ]

        for i, (code, meaning) in enumerate(decimal_data):
            ttk.Label(legend_frame, text=code, style='Legend.TLabel').grid(row=9 + i, column=0, sticky='ew')
            ttk.Label(legend_frame, text=meaning, style='Legend.TLabel').grid(row=9 + i, column=1, sticky='ew')

        legend_frame.grid_columnconfigure(0, weight=1)
        legend_frame.grid_columnconfigure(1, weight=1)

        # --- Action and Status Section ---
        
        # Process Button (Backup is no longer part of this action)
        self.process_button = ttk.Button(main_frame, text="Insert Row and Save Changes", command=self.process_insertion)
        self.process_button.grid(row=4, column=0, columnspan=2, pady=25)
        
        # Status Label
        self.status_text = tk.StringVar()
        # Display initial backup status in the main status area after GUI loads
        if self.initial_backup_status == "SUCCESS":
            self.status_text.set(f"Ready. Initial backup successful:\n{self.initial_backup_message}")
            self.status_label = ttk.Label(main_frame, textvariable=self.status_text, style='Success.TLabel', wraplength=600)
        else:
            self.status_text.set(f"Error. Initial backup failed:\n{self.initial_backup_message}.\nProceed with caution.")
            self.status_label = ttk.Label(main_frame, textvariable=self.status_text, style='Error.TLabel', wraplength=600)
            
        self.status_label.grid(row=5, column=0, columnspan=2, sticky='ew')
        
    # --- UI Helpers ---

    def browse_target_file(self):
        """Opens a file dialog to select the single target Excel file."""
        filetypes = [("Excel files", "*.xlsx")]
        filename = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=filetypes)
        if filename:
            self.target_file_path.set(filename)

    def create_backup(self, original_file):
        """
        Creates a timestamped backup of the original file using the required naming convention,
        including seconds to ensure unique file names on rapid runs.
        Returns (success_bool, backup_filepath_or_error_message)
        """
        # File existence check for safety
        if not os.path.exists(original_file):
            return False, f"Original file not found at '{original_file}' for backup."

        # Get the filename and extension (e.g., ('Movies and Shows', '.xlsx'))
        base, ext = os.path.splitext(original_file)
        
        # Generate timestamp string: ' - Backup yyyy mm dd hh mm ss'
        timestamp = datetime.now().strftime(" - Backup %Y %m %d %H %M %S")
        
        backup_file = f"{base}{timestamp}{ext}"

        try:
            # copy2 preserves metadata like modified time
            shutil.copy2(original_file, backup_file)
            return True, backup_file
        except Exception as e:
            # Handle potential file system errors (e.g., permission denied)
            return False, f"Failed to create backup: {type(e).__name__}: {e}"

    def copy_cell_properties(self, source_cell, target_cell, copy_number_format=True):
        """
        Copies style/formatting properties from source_cell to target_cell.
        If copy_number_format is False, the number_format is not transferred.
        """
        if source_cell.has_style:
            target_cell.style = source_cell.style
            
        # Only copy number format if explicitly requested. 
        # This is skipped for Column J to force a standard date format later.
        if copy_number_format and source_cell.number_format:
            target_cell.number_format = source_cell.number_format

    def copy_formulas(self, ws, source_row, target_row):
        """
        Copies the formula content and formatting from columns B, C, D, and E of 
        the source_row to the target_row, updating relative references.
        """
        # Columns to copy formulas from (1-based index: 2=B, 3=C, 4=D, 5=E)
        formula_cols = [2, 3, 4, 5] 
        
        # Define the row number strings for replacement
        old_row_str = str(source_row)
        new_row_str = str(target_row)
        
        # Columns that contain relative references that need to be updated (A and B)
        # Based on typical formula structures in this spreadsheet.
        columns_to_update = ['A', 'B'] 
        
        replacement_targets = []
        for col in columns_to_update:
            # Uppercase relative (e.g., A231 -> A232)
            replacement_targets.append((f"{col}{old_row_str}", f"{col}{new_row_str}"))
            # Lowercase relative (e.g., a231 -> a232)
            replacement_targets.append((f"{col.lower()}{old_row_str}", f"{col.lower()}{new_row_str}"))
        
        for col_idx in formula_cols:
            source_cell = ws.cell(row=source_row, column=col_idx)
            target_cell = ws.cell(row=target_row, column=col_idx)
            
            # Note: Cell properties (style, number_format) are copied in the main 
            # insert_row_to_excel loop, so we only handle the formula here.
            
            # Check if it's a formula and has a value
            if source_cell.data_type == 'f' and source_cell.value:
                formula_text = source_cell.value
                updated_formula = formula_text
                
                # Apply the defined replacements sequentially
                for old_ref, new_ref in replacement_targets:
                    updated_formula = updated_formula.replace(old_ref, new_ref)
                
                target_cell.value = updated_formula
            else:
                # If it's not a formula, copy the value directly
                target_cell.value = source_cell.value
                
    def _parse_coordinate(self, coord_str):
        """Manually splits an Excel coordinate string (e.g., 'A10') into (column_letter, row_number_int)."""
        col = ""
        row_str = ""
        for char in coord_str:
            if char.isalpha():
                col += char.upper()
            elif char.isdigit():
                row_str += char
        
        if not col or not row_str:
             # Fallback or error handling for invalid coordinates
             raise ValueError(f"Invalid Excel coordinate string: {coord_str}")

        return (col, int(row_str))

    # --- Core Logic ---

    def insert_row_to_excel(self, target_file, new_row_data):
        """
        Loads the workbook, appends the new row, and saves the changes back to 
        the target file (backup must be handled BEFORE this function is called).
        """
        try:
            # 1. Load the existing workbook, keeping formulas intact (data_only=False)
            wb = load_workbook(target_file, data_only=False)
            ws = wb.active # Use the active sheet

            # 2. Identify existing table (if any)
            table_name = None
            table_range = None
            table_style = None
            max_row = ws.max_row
            row_start_num = 1 # Default start row for table if found

            if ws.tables:
                table_name = list(ws.tables.keys())[0]
                table = ws.tables[table_name]
                table_range = table.ref
                
                if hasattr(table, 'tableStyleInfo'):
                    table_style = table.tableStyleInfo
                
                start_coord_str, end_coord_str = table_range.split(':')
                start_col_letter, row_start_num = self._parse_coordinate(start_coord_str)
                end_col_letter, max_row = self._parse_coordinate(end_coord_str)
            else:
                start_col_letter, end_col_letter = 'A', 'M' # Assume A:M range if no table

            # --- Calculate Insertion Points (Append Logic) ---
            
            last_existing_row = max_row
            
            if last_existing_row <= 1:
                # Case 1: Only header or empty sheet -> Insert data at row 2
                new_row_number = 2
                formula_source_row = 1 
            else:
                # Case 2: Data exists -> Append data after the last row
                new_row_number = last_existing_row + 1
                formula_source_row = last_existing_row
            
            # 3. Write the new input data and copy formulas/formatting
            
            # Copy formatting from the source row to the new appended row for all columns
            if formula_source_row >= 1: 
                for col_idx in range(1, 14): # Columns A through M (1 to 13)
                    source_cell = ws.cell(row=formula_source_row, column=col_idx)
                    target_cell = ws.cell(row=new_row_number, column=col_idx)
                    
                    # Column J (index 10) needs a standard date format, so we skip 
                    is_column_j = (col_idx == 10)
                    
                    self.copy_cell_properties(source_cell, target_cell, copy_number_format=not is_column_j)

                # Copy Formulas (Columns B, C, D, E) specifically
                if formula_source_row > 1:
                    self.copy_formulas(ws, formula_source_row, new_row_number)
            
            # Write input data (A, F-M)
            ws.cell(row=new_row_number, column=1, value=new_row_data[0]) # Column A
            
            # Column F to M (indices 6 to 13). 
            for i, value in enumerate(new_row_data[1:]): 
                col_idx = 6 + i
                target_cell = ws.cell(row=new_row_number, column=col_idx, value=value)
                
                # If this is Column J (index 10) and a date object was provided, 
                # explicitly set the standard 'dd-mmm-yyyy' format.
                if col_idx == 10 and isinstance(value, date): 
                    target_cell.number_format = 'dd-mmm-yyyy' 
            
            # 4. Update the Table Reference
            if table_name and table_range:
                new_max_row = new_row_number
                new_ref = f"{start_col_letter}{row_start_num}:{end_col_letter}{new_max_row}"
                
                del ws.tables[table_name]
                
                new_table = Table(displayName=table_name, ref=new_ref)
                if table_style:
                    new_table.tableStyleInfo = table_style
                
                ws.add_table(new_table)
                
            # 5. Save the modified workbook to the target file (overwriting the original)
            wb.save(target_file)
            
            message = f"✅ Successfully appended new row at row {new_row_number} and saved to:\n{target_file}"
            if table_name:
                message += f"\nNote: Updated table '{table_name}' range to {new_ref}."
            
            return True, message

        except FileNotFoundError:
            # If the file wasn't found during load, the validation in process_insertion should have caught it, 
            # but this is a final fail-safe.
            return False, f"Error: Target file not found at:\n{target_file}\nPlease ensure the file exists."
        except ValueError as e:
            return False, f"Error processing Excel coordinates: {e}"
        except Exception as e:
            # Provide specific error detail for debugging
            return False, f"An unexpected error occurred during file processing: {type(e).__name__}: {e}\n\nThis often happens if the Excel file is open (please close it) or the table is malformed."

    def process_insertion(self):
        """Validates inputs and initiates the Excel processing. Backup is handled on app load."""
        target_file = self.target_file_path.get().strip() 
        code_str = self.col_a.get().strip()
        title = self.col_f.get().strip()
        
        # --- 1. Validation checks ---
        if not all([target_file, code_str, title]):
            self.status_text.set("Error: Target File, Code (A), and Title (F) are mandatory fields.")
            self.status_label.config(style='Error.TLabel')
            return
            
        # --- 2. Data Validation and Type Conversion ---
        
        # 2a. Convert 'code' to float
        try:
            code = float(code_str)
        except ValueError:
            self.status_text.set("Error: Input Code (A) must be a valid number (integer or decimal, e.g., 1234 or 1234.5). Please correct the input.")
            self.status_label.config(style='Error.TLabel')
            return
            
        # Optional fields
        col_g = self.col_g.get().strip()
        col_h = self.col_h.get().strip()
        col_i = self.col_i.get().strip()
        col_j_str = self.col_j.get().strip() 
        col_k = self.col_k.get().strip()
        col_l = self.col_l.get().strip()
        col_m = self.col_m.get().strip()

        # 2b. Validate and Convert Date (Column J)
        date_j = col_j_str
        if col_j_str:
            date_formats = [
                '%d-%b-%Y', '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y',
            ]
            
            parsed_date = None
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(col_j_str, fmt).date() 
                    break 
                except ValueError:
                    continue 

            if parsed_date is None:
                self.status_text.set(f"Error: Date format in Column J ('{col_j_str}') is invalid. Please use a recognizable date format (e.g., 1/1/2025, 01/10/25, or 15-Mar-2020).")
                self.status_label.config(style='Error.TLabel')
                return
            
            date_j = parsed_date

        # Prepare data list
        new_row_data = [
            code, title, col_g, col_h, col_i, date_j, col_k, col_l, col_m
        ]

        # --- 3. Call Core Excel Logic ---
        self.status_text.set("Processing file... Please wait.")
        self.status_label.config(style='TLabel') 
        self.master.update() 

        success, message = self.insert_row_to_excel(target_file, new_row_data)

        # Update status message
        self.status_text.set(message)
        if success:
            self.status_label.config(style='Success.TLabel')
            # Clear input fields upon success
            self.col_a.set("")
            self.col_f.set("")
            self.col_g.set("")
            self.col_h.set("")
            self.col_j.set("")
            self.col_k.set("")
            self.col_l.set("")
            self.col_m.set("")
        else:
            self.status_label.config(style='Error.TLabel')


if __name__ == '__main__':
    # Check for openpyxl dependency
    try:
        from openpyxl import load_workbook
        import sys
        import shutil 
    except ImportError:
        root = tk.Tk()
        root.withdraw() 
        messagebox.showerror(
            "Missing Dependency",
            "The 'openpyxl' library is required to run this script.\n"
            "Please install it using your terminal:\n\n"
            "pip install openpyxl"
        )
        sys.exit(1)

    # Start the GUI application
    root = tk.Tk()
    app = ExcelRowInserterApp(root)
    root.mainloop()

